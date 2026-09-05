import os
import uuid
import shutil
from pathlib import Path

from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename

from pdf2docx import Converter
import pdfplumber
from openpyxl import Workbook

BASE = Path(__file__).resolve().parent
UPLOADS = BASE / "uploads"
OUTPUTS = BASE / "outputs"
UPLOADS.mkdir(exist_ok=True)
OUTPUTS.mkdir(exist_ok=True)

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

ALLOWED_PDF = {".pdf"}
ALLOWED_WORD = {".doc", ".docx"}

def cleanup(*paths):
    for p in paths:
        try:
            Path(p).unlink(missing_ok=True)
        except Exception:
            pass

@app.get("/")
def index():
    return render_template("index.html")

@app.post("/api/pdf-to-word")
def pdf_to_word():
    if "file" not in request.files:
        return jsonify(error="Please choose a PDF file."), 400

    f = request.files["file"]
    original = secure_filename(f.filename or "")
    if Path(original).suffix.lower() not in ALLOWED_PDF:
        return jsonify(error="Only PDF files are supported."), 400

    job = uuid.uuid4().hex
    pdf_path = UPLOADS / f"{job}.pdf"
    docx_path = OUTPUTS / f"{job}.docx"
    f.save(pdf_path)

    try:
        converter = Converter(str(pdf_path))
        converter.convert(str(docx_path), start=0, end=None)
        converter.close()

        if not docx_path.exists():
            raise RuntimeError("The PDF-to-Word engine did not create an output file.")

        return send_file(
            docx_path,
            as_attachment=True,
            download_name=f"{Path(original).stem}.docx",
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
    except Exception as e:
        return jsonify(error=f"PDF to Word failed: {e}"), 500
    finally:
        cleanup(pdf_path, docx_path)

@app.post("/api/pdf-to-excel")
def pdf_to_excel():
    if "file" not in request.files:
        return jsonify(error="Please choose a PDF file."), 400

    f = request.files["file"]
    original = secure_filename(f.filename or "")
    if Path(original).suffix.lower() not in ALLOWED_PDF:
        return jsonify(error="Only PDF files are supported."), 400

    job = uuid.uuid4().hex
    pdf_path = UPLOADS / f"{job}.pdf"
    xlsx_path = OUTPUTS / f"{job}.xlsx"
    f.save(pdf_path)

    try:
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        extracted = 0

        with pdfplumber.open(pdf_path) as pdf:
            for page_no, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                if not tables:
                    continue

                for table_no, table in enumerate(tables, start=1):
                    ws = wb.create_sheet(title=f"Page {page_no} Table {table_no}"[:31])
                    for row in table:
                        ws.append([(cell or "").strip() if isinstance(cell, str) else cell for cell in row])
                    extracted += 1

        if extracted == 0:
            # Still return a useful workbook containing page text when no table is detected.
            ws = wb.create_sheet("Extracted Text")
            with pdfplumber.open(pdf_path) as pdf:
                for page_no, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text() or ""
                    ws.cell(page_no, 1, f"Page {page_no}")
                    ws.cell(page_no, 2, text)

        wb.save(xlsx_path)

        return send_file(
            xlsx_path,
            as_attachment=True,
            download_name=f"{Path(original).stem}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        return jsonify(error=f"PDF to Excel failed: {e}"), 500
    finally:
        cleanup(pdf_path, xlsx_path)

@app.post("/api/word-to-pdf")
def word_to_pdf():
    # Optional companion endpoint. Requires LibreOffice installed on the server.
    if "file" not in request.files:
        return jsonify(error="Please choose a Word file."), 400

    f = request.files["file"]
    original = secure_filename(f.filename or "")
    if Path(original).suffix.lower() not in ALLOWED_WORD:
        return jsonify(error="Upload a DOC or DOCX file."), 400

    job = uuid.uuid4().hex
    input_path = UPLOADS / f"{job}{Path(original).suffix.lower()}"
    f.save(input_path)

    try:
        import subprocess
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", str(OUTPUTS), str(input_path)],
            check=True, capture_output=True, text=True, timeout=120
        )
        generated = OUTPUTS / f"{input_path.stem}.pdf"
        if not generated.exists():
            raise RuntimeError("LibreOffice did not create a PDF.")
        return send_file(generated, as_attachment=True,
                         download_name=f"{Path(original).stem}.pdf",
                         mimetype="application/pdf")
    except FileNotFoundError:
        return jsonify(error="Word to PDF requires LibreOffice on the server."), 501
    except Exception as e:
        return jsonify(error=f"Word to PDF failed: {e}"), 500
    finally:
        cleanup(input_path, OUTPUTS / f"{job}.pdf")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False)
